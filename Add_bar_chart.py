from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#load your table
file = load_workbook('pivot_table.xlsx')
sheet = file['Report']

min_column = file.active.min_column
max_column = file.active.max_column
min_row = file.active.min_row
max_row = file.active.max_row

barchart = BarChart()

#these data are particular column without titles
data = Reference(sheet,
          min_col=min_column+1,
          max_col=max_column,
          min_row=min_row,
          max_row=max_row,
)

# categories are particular row without titles
categories = Reference(sheet,
          min_col=min_column,
          max_col=min_column,
          min_row=min_row+1,
          max_row=max_row,
)

#adding your data to barchart
barchart.add_data(data)

#setting your categories to catogory
barchart.set_categories(categories)

#adding the barchart details to a particular cell
sheet.add_chart(barchart, 'B12')

#name a title
barchart.title = 'Trading chart'

# naming a style (eg. 1,2,3,4,5)
barchart.style = 5

#save the data into file
file.save('barchart.xlsx')
