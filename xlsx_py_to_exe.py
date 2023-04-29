from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os
import sys

application_path = os.path.dirname(sys.executable)

month = input("Enter the month: ")

#load your table
input_path = os.path.join(application_path, 'pivot_table.xlsx')
file = load_workbook(input_path)
sheet = file['Report']

min_column = file.active.min_column
max_column = file.active.max_column
min_row = file.active.min_row
max_row = file.active.max_row

barchart = BarChart()
#these data are particular column without titles
data = Reference(sheet,
          min_col=min_column+1,
          max_col=max_column,
          min_row=min_row,
          max_row=max_row,
)

# categories are particular row without titles
categories = Reference(sheet,
          min_col=min_column,
          max_col=min_column,
          min_row=min_row+1,
          max_row=max_row,
)

sheet['A1'] = 'Sales Report'
sheet['A2'] = month

#writes the sales report with size 20 and january size 10
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

output_path = os.path.join(application_path, f'report_{month}.xlsx')

file.save(f'report_{month}.xlsx')


#Follow this to convert py to exe file

""" first install "pip install pyinstaller" 
next pyinstaller --onefile --noconsole filename.py

IMPORTANT: to convert py to exe you have to delete 
        some modules in order to run those commands"""