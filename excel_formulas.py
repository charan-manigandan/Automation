from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


#load your table
file = load_workbook('barchart.xlsx')
sheet = file['Report']

min_column = file.active.min_column
max_column = file.active.max_column
min_row = file.active.min_row
max_row = file.active.max_row

# sheet['B9'] = '=SUM(B6,C7)'
# sheet['B9'].style = 'Currency'

#looping to calculate multiple rows and columns
for i in range(min_column+1, max_column+1):
    print(i)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

file.save('report.xlsx')